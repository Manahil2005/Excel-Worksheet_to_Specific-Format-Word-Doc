#!/usr/bin/env python3
"""
Watches a folder for changes to ANY .xlsx file (any name) and automatically
regenerates the matching Summary docx by calling generate_spi_summary.py.

It doesn't rely on the filename at all. Instead it opens the changed workbook,
reads the report date the same way generate_spi_summary.py already does
(Impact Combined sheet, row 2: "Current Date = DD/MM/YYYY, ..."), and names the
output Summary_DD_MM_YYYY.docx from that. So you can call your worksheet
anything -- Worksheet_final_v2.xlsx, thisweek.xlsx, whatever -- and it still
produces the correctly-named, correctly-dated summary.

If a workbook doesn't have that sheet/row (i.e. it's not an SPI worksheet at
all), it's skipped automatically so unrelated .xlsx files in the folder are
left alone.

USAGE:
    python3 watch_and_generate.py --dir "C:\\path\\to\\your\\SPI folder"

    (defaults to the current folder if --dir is omitted)

Leave this running in a terminal window while you work. Stop with Ctrl+C.
Requires openpyxl (already needed by generate_spi_summary.py) plus the
standard library -- nothing extra to install.
"""
import argparse
import os
import re
import subprocess
import sys
import time

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GENERATOR = os.path.join(SCRIPT_DIR, "generate_spi_summary.py")

DATE_RE = re.compile(r"Current Date\s*=\s*([\d/]+)")


def read_report_date(xlsx_path):
    """Returns 'DD_MM_YYYY' string read from inside the workbook, or None if this
    doesn't look like an SPI worksheet (missing sheet/row/pattern)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return None
    try:
        if "Impact Combined" not in wb.sheetnames:
            return None
        ws = wb["Impact Combined"]
        text = ws.cell(row=2, column=1).value
        if not isinstance(text, str):
            return None
        m = DATE_RE.search(text)
        if not m:
            return None
        d, mo, y = m.group(1).strip().split("/")
        return f"{d.zfill(2)}_{mo.zfill(2)}_{y}"
    except Exception:
        return None
    finally:
        wb.close()


def is_file_stable(path, checks=2, interval=0.5):
    """Waits until file size stops changing, so we don't read a half-saved xlsx."""
    last = -1
    stable_count = 0
    for _ in range(20):
        try:
            size = os.path.getsize(path)
        except OSError:
            return False
        if size == last:
            stable_count += 1
            if stable_count >= checks:
                return True
        else:
            stable_count = 0
        last = size
        time.sleep(interval)
    return True


def regenerate(xlsx_path, watch_dir):
    fname = os.path.basename(xlsx_path)

    if not is_file_stable(xlsx_path):
        print(f"[skip] {fname} still being written, will catch it next change")
        return

    date_str = read_report_date(xlsx_path)
    if date_str is None:
        print(f"[skip] {fname} doesn't look like an SPI worksheet (no 'Impact Combined' date row) -- ignoring")
        return

    out_name = f"Summary_{date_str}.docx"
    out_path = os.path.join(watch_dir, out_name)

    print(f"[change detected] {fname} (report date {date_str}) -> regenerating {out_name} ...")
    result = subprocess.run(
        [sys.executable, GENERATOR, "--excel", xlsx_path, "--output", out_path],
        capture_output=True, text=True,
    )
    if result.returncode == 0:
        print(f"[done] {out_name} updated")
    else:
        print(f"[error] generation failed for {fname}:\n{result.stderr}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=".", help="Folder to watch for .xlsx files")
    ap.add_argument("--interval", type=float, default=2.0, help="Poll interval in seconds")
    args = ap.parse_args()

    watch_dir = os.path.abspath(args.dir)
    if not os.path.isdir(watch_dir):
        sys.exit(f"Not a folder: {watch_dir}")
    if not os.path.exists(GENERATOR):
        sys.exit(f"Can't find generate_spi_summary.py next to this watcher at {GENERATOR}")

    print(f"Watching {watch_dir} for any .xlsx changes (Ctrl+C to stop)...")
    last_mtimes = {}

    try:
        while True:
            for fname in os.listdir(watch_dir):
                lower = fname.lower()
                # skip non-xlsx files, Excel's own temp lock files, and anything
                # that isn't a real spreadsheet
                if not lower.endswith(".xlsx") or fname.startswith("~$"):
                    continue
                fpath = os.path.join(watch_dir, fname)
                try:
                    mtime = os.path.getmtime(fpath)
                except OSError:
                    continue
                if last_mtimes.get(fname) != mtime:
                    last_mtimes[fname] = mtime
                    regenerate(fpath, watch_dir)
            time.sleep(args.interval)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()